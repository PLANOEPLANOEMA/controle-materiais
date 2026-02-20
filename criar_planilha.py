import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import os

# Dados dos materiais de empréstimo
dados = [
    {"empresa": "Felipe Agosti", "material": "Pontalete",          "data": "28/01/2026", "quantidade": 20},
    {"empresa": "Felipe Agosti", "material": "Chapa de Madeirite OSB", "data": "28/01/2026", "quantidade": 20},
    {"empresa": "Felipe Agosti", "material": "Pontalete",          "data": "12/02/2026", "quantidade": 150},
    {"empresa": "Felipe Agosti", "material": "Tábua 30cm",         "data": "12/02/2026", "quantidade": 50},
    {"empresa": "Murilo",        "material": "Pontalete",          "data": "04/02/2026", "quantidade": 220},
    {"empresa": "Spera Urban",   "material": "Pontalete",          "data": "03/02/2025", "quantidade": 100},
    {"empresa": "Plano Purus",   "material": "Tábua de 30cm",      "data": "06/02/2026", "quantidade": 100},
]

# Criar workbook
wb = openpyxl.Workbook()

# ─────────────────────────────────────────────
# ABA 1 – REGISTRO COMPLETO
# ─────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Registro Completo"

# Paleta de cores
COR_TITULO_BG    = "1F3864"   # azul escuro
COR_TITULO_FG    = "FFFFFF"
COR_HEADER_BG    = "2E75B6"   # azul médio
COR_HEADER_FG    = "FFFFFF"
COR_LINHA_PAR    = "DDEEFF"   # azul muito claro
COR_LINHA_IMPAR  = "FFFFFF"
COR_BORDA        = "2E75B6"
COR_TOTAL_BG     = "1F3864"
COR_TOTAL_FG     = "FFFFFF"

borda = Border(
    left=Side(style="thin", color=COR_BORDA),
    right=Side(style="thin", color=COR_BORDA),
    top=Side(style="thin", color=COR_BORDA),
    bottom=Side(style="thin", color=COR_BORDA),
)

# Título principal (linha 1, colunas A-F)
ws1.merge_cells("A1:F1")
titulo = ws1["A1"]
titulo.value = "CONTROLE DE MATERIAIS DE EMPRÉSTIMO"
titulo.font = Font(name="Calibri", bold=True, size=16, color=COR_TITULO_FG)
titulo.fill = PatternFill("solid", fgColor=COR_TITULO_BG)
titulo.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

# Subtítulo (linha 2)
ws1.merge_cells("A2:F2")
sub = ws1["A2"]
sub.value = "Gerado automaticamente – todos os débitos em aberto"
sub.font = Font(name="Calibri", italic=True, size=10, color="7F7F7F")
sub.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 18

# Cabeçalhos (linha 3)
cabecalhos = ["#", "Empresa", "Material", "Data do Empréstimo", "Quantidade", "Status"]
for col, cab in enumerate(cabecalhos, start=1):
    cell = ws1.cell(row=3, column=col, value=cab)
    cell.font = Font(name="Calibri", bold=True, size=11, color=COR_HEADER_FG)
    cell.fill = PatternFill("solid", fgColor=COR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = borda
ws1.row_dimensions[3].height = 28

# Dados
for i, d in enumerate(dados, start=1):
    row = i + 3
    fill_bg = COR_LINHA_PAR if i % 2 == 0 else COR_LINHA_IMPAR
    fill = PatternFill("solid", fgColor=fill_bg)

    valores = [i, d["empresa"], d["material"], d["data"], d["quantidade"], "Em aberto"]
    for col, val in enumerate(valores, start=1):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = borda
        cell.alignment = Alignment(horizontal="center" if col in (1, 4, 5, 6) else "left",
                                   vertical="center")
        cell.font = Font(name="Calibri", size=11)
        if col == 6:
            cell.font = Font(name="Calibri", size=11, bold=True, color="C00000")
    ws1.row_dimensions[row].height = 22

# Linha de total
total_row = len(dados) + 4
ws1.merge_cells(f"A{total_row}:D{total_row}")
total_label = ws1[f"A{total_row}"]
total_label.value = "TOTAL GERAL DE ITENS"
total_label.font = Font(name="Calibri", bold=True, size=11, color=COR_TOTAL_FG)
total_label.fill = PatternFill("solid", fgColor=COR_TOTAL_BG)
total_label.alignment = Alignment(horizontal="right", vertical="center")
total_label.border = borda

total_val = ws1.cell(row=total_row, column=5, value=sum(d["quantidade"] for d in dados))
total_val.font = Font(name="Calibri", bold=True, size=12, color=COR_TOTAL_FG)
total_val.fill = PatternFill("solid", fgColor=COR_TOTAL_BG)
total_val.alignment = Alignment(horizontal="center", vertical="center")
total_val.border = borda

ws1.cell(row=total_row, column=6).fill = PatternFill("solid", fgColor=COR_TOTAL_BG)
ws1.cell(row=total_row, column=6).border = borda
ws1.row_dimensions[total_row].height = 26

# Larguras das colunas
larguras = [5, 22, 28, 22, 14, 14]
for col, larg in enumerate(larguras, start=1):
    ws1.column_dimensions[get_column_letter(col)].width = larg

# Congelar painel abaixo do cabeçalho
ws1.freeze_panes = "A4"

# ─────────────────────────────────────────────
# ABA 2 – RESUMO POR EMPRESA
# ─────────────────────────────────────────────
ws2 = wb.create_sheet("Resumo por Empresa")

# Agregar por empresa
resumo_empresa = {}
for d in dados:
    emp = d["empresa"]
    resumo_empresa[emp] = resumo_empresa.get(emp, 0) + d["quantidade"]

# Título
ws2.merge_cells("A1:C1")
t2 = ws2["A1"]
t2.value = "RESUMO POR EMPRESA"
t2.font = Font(name="Calibri", bold=True, size=14, color=COR_TITULO_FG)
t2.fill = PatternFill("solid", fgColor=COR_TITULO_BG)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

# Cabeçalhos
for col, cab in enumerate(["Empresa", "Total de Itens", "% do Total"], start=1):
    cell = ws2.cell(row=2, column=col, value=cab)
    cell.font = Font(name="Calibri", bold=True, size=11, color=COR_HEADER_FG)
    cell.fill = PatternFill("solid", fgColor=COR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = borda
ws2.row_dimensions[2].height = 26

total_geral = sum(resumo_empresa.values())
for i, (emp, qtd) in enumerate(sorted(resumo_empresa.items(), key=lambda x: -x[1]), start=1):
    row = i + 2
    fill_bg = COR_LINHA_PAR if i % 2 == 0 else COR_LINHA_IMPAR
    fill = PatternFill("solid", fgColor=fill_bg)

    ws2.cell(row=row, column=1, value=emp).fill = fill
    ws2.cell(row=row, column=1).border = borda
    ws2.cell(row=row, column=1).font = Font(name="Calibri", size=11)
    ws2.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    ws2.cell(row=row, column=2, value=qtd).fill = fill
    ws2.cell(row=row, column=2).border = borda
    ws2.cell(row=row, column=2).font = Font(name="Calibri", size=11, bold=True)
    ws2.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")

    pct = round(qtd / total_geral * 100, 1)
    ws2.cell(row=row, column=3, value=f"{pct}%").fill = fill
    ws2.cell(row=row, column=3).border = borda
    ws2.cell(row=row, column=3).font = Font(name="Calibri", size=11)
    ws2.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[row].height = 22

# Total
total_row2 = len(resumo_empresa) + 3
ws2.merge_cells(f"A{total_row2}:B{total_row2}")
tc = ws2[f"A{total_row2}"]
tc.value = f"TOTAL GERAL: {total_geral} itens"
tc.font = Font(name="Calibri", bold=True, size=11, color=COR_TOTAL_FG)
tc.fill = PatternFill("solid", fgColor=COR_TOTAL_BG)
tc.alignment = Alignment(horizontal="center", vertical="center")
tc.border = borda
ws2.cell(row=total_row2, column=3).fill = PatternFill("solid", fgColor=COR_TOTAL_BG)
ws2.cell(row=total_row2, column=3).border = borda
ws2.row_dimensions[total_row2].height = 26

ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 14

# ─────────────────────────────────────────────
# ABA 3 – RESUMO POR MATERIAL
# ─────────────────────────────────────────────
ws3 = wb.create_sheet("Resumo por Material")

resumo_material = {}
for d in dados:
    mat = d["material"]
    resumo_material[mat] = resumo_material.get(mat, 0) + d["quantidade"]

ws3.merge_cells("A1:C1")
t3 = ws3["A1"]
t3.value = "RESUMO POR MATERIAL"
t3.font = Font(name="Calibri", bold=True, size=14, color=COR_TITULO_FG)
t3.fill = PatternFill("solid", fgColor=COR_TITULO_BG)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 32

for col, cab in enumerate(["Material", "Total de Itens", "% do Total"], start=1):
    cell = ws3.cell(row=2, column=col, value=cab)
    cell.font = Font(name="Calibri", bold=True, size=11, color=COR_HEADER_FG)
    cell.fill = PatternFill("solid", fgColor=COR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = borda
ws3.row_dimensions[2].height = 26

for i, (mat, qtd) in enumerate(sorted(resumo_material.items(), key=lambda x: -x[1]), start=1):
    row = i + 2
    fill_bg = COR_LINHA_PAR if i % 2 == 0 else COR_LINHA_IMPAR
    fill = PatternFill("solid", fgColor=fill_bg)

    ws3.cell(row=row, column=1, value=mat).fill = fill
    ws3.cell(row=row, column=1).border = borda
    ws3.cell(row=row, column=1).font = Font(name="Calibri", size=11)
    ws3.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    ws3.cell(row=row, column=2, value=qtd).fill = fill
    ws3.cell(row=row, column=2).border = borda
    ws3.cell(row=row, column=2).font = Font(name="Calibri", size=11, bold=True)
    ws3.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")

    pct = round(qtd / total_geral * 100, 1)
    ws3.cell(row=row, column=3, value=f"{pct}%").fill = fill
    ws3.cell(row=row, column=3).border = borda
    ws3.cell(row=row, column=3).font = Font(name="Calibri", size=11)
    ws3.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[row].height = 22

total_row3 = len(resumo_material) + 3
ws3.merge_cells(f"A{total_row3}:B{total_row3}")
tc3 = ws3[f"A{total_row3}"]
tc3.value = f"TOTAL GERAL: {total_geral} itens"
tc3.font = Font(name="Calibri", bold=True, size=11, color=COR_TOTAL_FG)
tc3.fill = PatternFill("solid", fgColor=COR_TOTAL_BG)
tc3.alignment = Alignment(horizontal="center", vertical="center")
tc3.border = borda
ws3.cell(row=total_row3, column=3).fill = PatternFill("solid", fgColor=COR_TOTAL_BG)
ws3.cell(row=total_row3, column=3).border = borda
ws3.row_dimensions[total_row3].height = 26

ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 14

# Salvar
os.makedirs("/home/ubuntu/controle_emprestimos", exist_ok=True)
caminho = "/home/ubuntu/controle_emprestimos/Controle_Materiais_Emprestimo.xlsx"
wb.save(caminho)
print(f"Planilha salva em: {caminho}")
